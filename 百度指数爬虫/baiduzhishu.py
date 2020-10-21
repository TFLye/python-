import os
import json
from time import sleep
import openpyxl
import requests
import datetime
import xlrd
import smtplib
from email.mime.text import MIMEText
from email.header import Header

# 加载省份和城市的json数据,以及对应关系
# e.g. '1': '济南'
provinces = json.loads(open('province.json',encoding='utf-8').read())
city = json.loads(open('city.json',encoding='utf-8').read())
pro_city = json.loads(open('pro_city.json').read())
# 配置相关
START_DATE = "2011-01-03"
END_DATE = "2019-12-30"
NROWS = 1324 # 股票信息execl总行数
FILE_NAME = '数据.xlsx'
root_path = 'alldata\\'
### 请爬取数据前将自己的cookie放这
cookie = ''
who_spider = '我是你的爬虫~'  ### 出错时发送邮件提醒
# 将数据按格式存入表格
# data_list : 城市  证券代码  证券简称   周数据量
#             city  code     name    week_data   time

# 保存数据，省份，城市下标，要写的行数，要写入的数据

# 百度指数data解密函数
def decrypt(password,data):
    arr = list(password)
    dataArr = list(data)
    objPass = {}
    notInNames = []
    for i in range(0,int(len(arr)/2)):
        objPass[arr[i]] = arr[int(len(arr)/2)+i]
        i+=1
    # 数据解密转换
    for i in range(0,len(data)):
        notInNames.append(objPass[dataArr[i]])
    return ''.join(notInNames)
def send_email(message):
    # 发信方的信息：发信邮箱，QQ 邮箱授权码
    from_addr = '1234546@qq.com'
    password = ''
    # 收信方邮箱
    to_addr = ['123456@qq.com','456789@qq.com']
    # 发信服务器
    smtp_server = 'smtp.qq.com'
    # 邮箱正文内容，第一个参数为内容，第二个参数为格式(plain 为纯文本)，第三个参数为编码
    msg = MIMEText(message,'plain','utf-8')
    # 邮件头信息
    msg['From'] = Header(from_addr)
    msg['To'] = Header(','.join(to_addr))
    msg['Subject'] = "百度搜索指数爬虫^-^"
    # 开启发信服务，这里使用的是加密传输
    server = smtplib.SMTP_SSL(smtp_server)
    server.connect(smtp_server,465)
    # 登录发信邮箱
    server.login(from_addr, password)
    # 发送邮件
    server.sendmail(from_addr, to_addr, msg.as_string())
    # 关闭服务器
    server.quit()

# 获取PC移动数据,area,地区编号，word,股票信息,city_idx,city在该省的位置,pro_name,省名,word_idx,股票下标
# 返回解密后的数据
def get_data(area,word,city_idx,pro_name,word_idx):
    data_url_temp = f"http://index.baidu.com/api/SearchApi/index?area={area}&word=[[%7B%22name%22:%22{word[0]}%22,%22wordType%22:1%7D,%7B%22name%22:%22{word[1]}%22,%22wordType%22:1%7D]]&startDate={START_DATE}&endDate={END_DATE}"
    data_header = {
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36 Edg/85.0.564.51',
        'Cookie':cookie,
    }
    # 请求数据
    data_r = requests.get(url=data_url_temp,headers=data_header)
    # 接收返回的所有数据
    data_dic = json.loads(data_r.text)['data']
    try:
        # 取出请求密码表passowrd的密钥uniqid
        uniqid = data_dic['uniqid']
    except Exception as e:
        send_email(f"{who_spider}\n爬虫中止:当前已爬到{pro_name}省,{city[area]}市,下标为{city_idx},第{word_idx}支股票!!!\n错误信息:{e}\n")

    # 取出PC端相关数据
    data = data_dic['userIndexes'][0]['pc']['data']


    # 根据uniqid请求password
    password_url = f"https://index.baidu.com/Interface/ptbk?uniqid={uniqid}"
    password_header = {
        'Host':'index.baidu.com',
        'Referer':'https://index.baidu.com/v2/main/index.html',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36 Edg/85.0.564.51',
        'Cookie':cookie,
    }
    # 请求密码表
    password_r = requests.get(url=password_url,headers=password_header)
    # 取出密码表password
    password = json.loads(password_r.text)['data']

    # print(f'data: {data}')
    # 输出解密后的数据
    real_data = decrypt(password,data)
    # print(f'real_data: {real_data}')
    return real_data

# 获取股票信息,row,读取excel中第row行的数据,
# 返回word列表
# e.g. ["000002","万科a"]
def get_word(row):
    # 读取excel文件
    word_data = xlrd.open_workbook("关键词.xls")
    sheet1 = word_data.sheets()[0]
    word = []
    word.append(sheet1.cell(row-1,2).value)
    word.append(sheet1.cell(row-1,0).value)
    # print(f"word:{word}")
    return word

# 完整处理函数
# start_row,从给定的起始word的行数;pro_name,city所处省份名称
# city_idx,city所处省份的位置;word_idx,第几支股票
# 爬虫中断所要提示数据：
#   1.当前的word及excel行数，
#   2.当前的city及省份
#   3.当前存到文件的哪个工作簿哪一行
def deal(pro_name,city_idx,word_idx):
    ###################################
    def save_data(data, province, city):
        # 不存在该文件则创建
        save_path = root_path+province
        city_data_path = save_path+'\\'+city+'.xlsx'
        if not os.path.exists(save_path):
            os.makedirs(save_path)
            print(f'新省份{province}创建成功')
        if not os.path.exists(city_data_path):
            # 只能这样创建，open创建的xlsx，openpyxl打不开！
            d = openpyxl.Workbook()  # 新建工作簿
            # d.create_sheet('test')  # 添加页
            d.save(city_data_path)
        # 加载xlsx文件
        wb = openpyxl.load_workbook(city_data_path)
        sheets = wb.sheetnames  # 获取所有工作簿的名称
        # 不存在则创建该省的工作簿
        # if city not in sheets:
        #     wb.create_sheet(city)
        sheet = wb[sheets[0]]  # 获取该省工作簿
        sheet.title = city
        delay = datetime.timedelta(days=7)  # 七天延迟
        w_row = sheet.max_row+1 # 下次开始保存股票信息的起始行数
        # 遍历数据字典，key:股票下标，value,真正数据
        for key in data.keys():
            word = data[key][0]
            week_data = data[key][1].split(',')  # 周数据list
            time = datetime.datetime.strptime(START_DATE, "%Y-%m-%d")  # 起始时间xi
            # 开始写入数据
            for i in range(len(week_data)):
                sheet.cell(w_row, 1).value = city  # 城市名称
                sheet.cell(w_row, 2).value = word[0]  # 股票代号
                sheet.cell(w_row, 3).value = word[1]  # 股票简称
                # 周数据
                if len(week_data[i]) == 0 : sheet.cell(w_row, 4).value = '0'
                else : sheet.cell(w_row, 4).value = week_data[i]
                sheet.cell(w_row, 5).value = str(time).split(' ')[0]  # 开始时间
                time += delay
                w_row += 1
                print(f'写入{city},{word[0]},{word[1]},{week_data[i]},{time}')
        wb.save(city_data_path)
        print('存入成功~')
        ###################################
    flag = False
    # 遍历所有省下的市,key,省名,value,城市代码
    for key,value in pro_city.items():
        # if key == "河北":
        #     return
        for c in value:
            # 检查是否是指定开始的地方
            # print(f"value: {value}   key: {key}  pro_name:{pro_name}   city_idx:{city_idx} value.index(c): {value.index(c)+1}")
            if (pro_name == key and city_idx == str((value.index(c)+1))) or flag:
                # 进入到指定位置（第一次）
                print("进入if")
                if not flag: idx = word_idx
                else : idx = 1
                interval = 1 # 每100次休息2s
                # 2.0更新
                # 每100次保存一下数据
                data_cnt = 0 # 每100次请求保存一下
                real_data = {} # 获取的真正数据,key:股票下标，value,真正数据

                # 开始遍历该城市的所有股票
                while idx <= NROWS :
                    if interval == 100:
                        print('人家爬累了,休息一会~')
                        # 休眠
                        sleep(2)
                        interval = 0
                    # 获取股票数据
                    word = get_word(idx)
                    # 获取城市代码
                    area = c
                    print(f"当前城市为{city[c]}, 正在爬取第{idx}支股票 {word[1]}, 正在爬爬爬~")
                    # 爬取数据
                    real_data[idx] = []
                    real_data[idx].append(word)
                    real_data[idx].append(get_data(area,word,city_idx,key,word_idx))
                    data_cnt += 1
                    # 第一百次，开始保存
                    if data_cnt == 100 or idx == NROWS:
                        data_cnt = 0;
                        # 保存数据
                        save_data(real_data,key,city[c])
                        real_data.clear()
                        print("保存成功！！！")
                    print("爬取成功！")
                    idx += 1
                    interval +=1
                flag = True

if __name__ == '__main__':
    ### 参数意义
    ### 省份名称 当前省的第n个市 第n只股票
    deal("内蒙古","6",1)
